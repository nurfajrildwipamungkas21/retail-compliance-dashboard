import os
import io
import base64
import textwrap
import logging
import sys
import re
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio  # for offline HTML generation
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Helpers for sample detection and Parquet fallback
def locate_sample() -> Path | None:
    """
    Locate a sample file under `sample_data` directory. It will prefer
    compressed CSV (.csv.gz), then plain CSV, then Parquet. If no file is found,
    returns None.
    """
    for p in [
        Path("sample_data/online_retail_sample.csv.gz"),
        Path("sample_data/online_retail_sample.csv"),
        Path("sample_data/online_retail_sample.parquet"),
    ]:
        if p.exists():
            return p
    return None

# Detect availability of pyarrow for Parquet read/write
HAVE_PARQUET = False
try:
    import pyarrow  # type: ignore
    HAVE_PARQUET = True
except Exception:
    HAVE_PARQUET = False

# CSV fallback paths for saving results when Parquet is not available
LAST_CLASSIFIED_CSV = "demo_last_classified.csv.gz"
LAST_REPORT_CSV = "demo_last_report.csv.gz"

# Set Streamlit page configuration
st.set_page_config(
    page_title="Compliance Intelligence Dashboard",
    page_icon="📊",
    layout="wide"
)
logging.getLogger().setLevel(logging.INFO)

# =========================
# THEME: Helper & CSS injector
# =========================
# Pilihan: "light" | "dark" | "auto"
THEME_KEY = "ui_theme_base"

def _get_theme_base() -> str:
    sel = st.session_state.get(THEME_KEY, "auto")
    if sel in ("light", "dark"):
        return sel
    # "auto" -> ikut setting Streamlit (user/browser)
    try:
        base = st.get_option("theme.base")
        return (base or "light").lower()
    except Exception:
        return "light"


def get_theme_palette(base: str) -> dict:
    if base == "dark":
        return dict(
            base="dark",
            bg="#0e1117",
            text="#e5e7eb",
            card="#111827",
            border="#1f2937",
            header="#f3f4f6",
            plotly_template="plotly_dark",
        )
    return dict(
        base="light",
        bg="#ffffff",
        text="#111827",
        card="#ffffff",
        border="#e5e7eb",
        header="#0f172a",
        plotly_template="plotly_white",
    )


def apply_theme():
    """
    Set template Plotly and inject global CSS per user theme choice.
    This extended version styles the entire app, including the top header,
    sidebar, file uploader, buttons, alerts, and dataframes to ensure
    consistency in dark mode.
    """
    base = _get_theme_base()
    pal = get_theme_palette(base)
    # set template plotly, simpan palette di session_state
    pio.templates.default = pal["plotly_template"]
    st.session_state["ui_palette"] = pal
    st.session_state["ui_text_color"] = pal["text"]
    # preserve resolved theme base for downstream use
    st.session_state["ui_theme_base_resolved"] = base

    st.markdown(f"""
    <style>
      /* Define CSS custom properties for dynamic theming */
      :root {{
        --ui-bg-color: {pal['bg']};
        --ui-text-color: {pal['text']};
        --ui-header-color: {pal['header']};
        --ui-card-color: {pal['card']};
        --ui-border-color: {pal['border']};
      }}

      /* Base background for app */
      html, body {{
        background: var(--ui-bg-color);
      }}
      .stApp {{
        background: var(--ui-bg-color);
      }}
      /* Ensure top header strip matches theme */
      header[data-testid="stHeader"] {{
        background: var(--ui-bg-color) !important;
        border-bottom: 1px solid var(--ui-border-color);
      }}
      /* Toolbar and header contents */
      [data-testid="stToolbar"],
      header[data-testid="stHeader"] * {{
        background: transparent !important;
        color: var(--ui-text-color) !important;
      }}
      /* App view and main container backgrounds */
      [data-testid="stAppViewContainer"],
      [data-testid="stMain"] {{
        background: var(--ui-bg-color) !important;
      }}

      /* Headings styling */
      h1, h2, h3, h4, h5, h6 {{
        color: var(--ui-header-color);
      }}

      /* Metrics styling */
      div[data-testid="stMetric"] label {{
        color: var(--ui-text-color) !important;
      }}
      div[data-testid="stMetricValue"] {{
        color: var(--ui-header-color) !important;
      }}

      /* Caption components: ensure they remain visible on dark backgrounds */
      .stCaption, .stCaption p {{
        color: var(--ui-text-color);
        opacity: 0.75;
      }}

      /* Markdown paragraphs and list items */
      div[data-testid="stMarkdownContainer"] p,
      div[data-testid="stMarkdownContainer"] li {{
        color: var(--ui-text-color);
      }}

      /* Emphasised text within markdown */
      div[data-testid="stMarkdownContainer"] strong {{
        color: var(--ui-header-color);
      }}

      /* Generic card styling */
      .card-like {{
        background: var(--ui-card-color);
        border: 1px solid var(--ui-border-color);
        border-radius: 12px;
        padding: 16px;
      }}

      /* Sidebar: background and text color */
      [data-testid="stSidebar"] > div:first-child {{
        background: var(--ui-bg-color);
        border-right: 1px solid var(--ui-border-color);
        color: var(--ui-text-color);
      }}
      [data-testid="stSidebar"] label,
      [data-testid="stSidebar"] span,
      [data-testid="stSidebar"] p {{
        color: var(--ui-text-color);
      }}

      /* =========================
         File uploader dark-mode
         ========================= */
      /* Wrapper: ensure no leftover white area under dropzone */
      [data-testid="stFileUploader"] {{
        background: var(--ui-bg-color) !important;
      }}
      /* Dropzone area */
      [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"],
      [data-testid="stFileUploader"] > div > div {{
        background: var(--ui-card-color) !important;
        border: 1px dashed var(--ui-border-color) !important;
        color: var(--ui-text-color) !important;
      }}
      /* Descriptive text and labels inside uploader */
      [data-testid="stFileUploader"] p,
      [data-testid="stFileUploader"] small,
      [data-testid="stFileUploader"] span,
      [data-testid="stFileUploader"] label {{
        color: var(--ui-text-color) !important;
      }}
      /* Browse files button styling */
      [data-testid="stFileUploader"] button {{
        background: var(--ui-card-color) !important;
        color: var(--ui-header-color) !important;
        border: 1px solid var(--ui-border-color) !important;
        border-radius: 8px !important;
      }}
      [data-testid="stFileUploader"] button:hover {{
        filter: brightness(1.08);
      }}
      /* Hover state for dropzone (dragging over) */
      [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"]:hover {{
        background: rgba(255,255,255,0.03) !important;
      }}
      /* File list preview area */
      [data-testid="stFileUploader"] [role="list"],
      [data-testid="stFileUploader"] [role="list"] * {{
        background: transparent !important;
        color: var(--ui-text-color) !important;
      }}

      /* Buttons: standard and download */
      button {{
        background: var(--ui-card-color) !important;
        color: var(--ui-header-color) !important;
        border: 1px solid var(--ui-border-color) !important;
        border-radius: 8px !important;
        padding: 0.4rem 0.75rem !important;
      }}
      button:hover {{
        filter: brightness(1.1);
      }}

      /* Badge 'Pipeline selesai' */
      [data-testid="stAlert"] {{
        background: #193d3d !important;
        border: 1px solid #0f2f2f !important;
        color: #c6f6f3 !important;
      }}

      /* DataFrame and table styling */
      .stDataFrameContainer, .stTable {{
        background: var(--ui-card-color);
        border-radius: 8px;
        border: 1px solid var(--ui-border-color);
        color: var(--ui-text-color);
      }}

      /* Ensure the top container uses app background in dark mode */
      .block-container {{
        background: var(--ui-bg-color);
      }}
    </style>
    """, unsafe_allow_html=True)


# =========================
# PERSISTENCE
# =========================
# File paths where the last pipeline results will be stored. These files are
# created after the user runs the pipeline or after the default sample is
# processed. On subsequent loads, if these files exist, their contents are
# loaded and displayed automatically.
LAST_CLASSIFIED_PATH = "demo_last_classified.parquet"
LAST_REPORT_PATH = "demo_last_report.parquet"

# =========================
# API KEY GEMINI (DEMO)
# =========================
# Fallback: use environment variable GEMINI_API_KEY if provided; otherwise
# fallback to this demo key. For production usage, consider storing the
# key securely in secrets and not hard-coding it here.
DEMO_GEMINI_KEY = "AIzaSyCXQQlDnENBgfFXsmDj86we_UxigK0TlgM"


def get_gemini_key() -> str:
    """Return API key from environment or fallback to demo key."""
    env_key = os.getenv("GEMINI_API_KEY", "").strip()
    return env_key if env_key else DEMO_GEMINI_KEY


# Initialize google-generativeai module if available
ENABLE_AI = True
try:
    import google.generativeai as genai  # type: ignore
    genai.configure(api_key=get_gemini_key())
except Exception as e:
    ENABLE_AI = False
    logging.warning(f"google-generativeai tidak aktif: {e}")


# =========================
# KONFIGURASI & KONSTANTA
# =========================

def normalize_col(name: str) -> str:
    """Normalize column names for matching against target columns."""
    return str(name).lower().replace(" ", "").replace("_", "")


TARGET_COLUMNS = {
    normalize_col(c)
    for c in [
        "UnitPrice",
        "CustomerID",
        "Country",
        "Quantity",
        "Description",
        "StockCode",
        "InvoiceNo",
        "InvoiceDate",
    ]
}

MIN_MATCH_SCORE = 5
VALID_DATA_THRESHOLD = 6


# =========================
# UTIL: EXCEL BYTES
# =========================

def df_to_excel_bytes(df: pd.DataFrame, title: str, summary_data: dict) -> bytes:
    """
    Convert a pandas DataFrame into an Excel file stored in memory with
    formatting, title, summary, dynamic column widths, and zebra striping.

    Returns the raw bytes of the Excel file for use in Streamlit download
    buttons.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write title
    ws.merge_cells("A1:E1")
    ws["A1"].value = title
    ws["A1"].font = Font(size=16, bold=True, color="0B4F6C")

    # Write summary below the title
    row_num = 3
    for k, v in summary_data.items():
        ws.cell(row=row_num, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=v)
        row_num += 1

    start_row = row_num + 2  # leave a blank row before data

    # Write DataFrame to worksheet
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start_row
    ):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    for cell in ws[start_row]:
        cell.font = header_font
        cell.fill = header_fill

    # Create table with zebra striping
    if df.shape[0] > 0 and df.shape[1] > 0:
        ref = f"A{start_row}:{get_column_letter(df.shape[1])}{start_row + df.shape[0]}"
        table = Table(displayName="DataTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# PIPELINE: LOADING & QC
# =========================

def try_read_any(path: Path) -> pd.DataFrame:
    """
    Read a file regardless of its extension (CSV, compressed CSV, Excel, or Parquet).

    If the file has a ``.csv.gz`` suffix, ``pandas.read_csv`` will automatically
    detect the gzip compression. Parquet files are only read if the optional
    dependency ``pyarrow`` is available; otherwise, a ``ValueError`` will be
    raised so that callers can handle missing Parquet support gracefully.
    """
    suffix = path.suffix.lower()
    name = path.name.lower()
    # Excel formats
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    # Compressed or plain CSV
    if suffix == ".csv" or name.endswith(".csv.gz"):
        try:
            return pd.read_csv(path)
        except UnicodeDecodeError:
            # fallback encoding for some European/Asian CSV files
            return pd.read_csv(path, encoding="latin1")
    # Parquet (requires pyarrow)
    if suffix == ".parquet":
        if HAVE_PARQUET:
            return pd.read_parquet(path)
        # If Parquet support isn't available, signal error to caller
        raise ValueError("Parquet support is not available (pyarrow not installed)")
    raise ValueError(f"Ekstensi tidak didukung: {path.suffix}")


def guess_best_candidate(base: Path) -> Path | None:
    """
    Search for the best candidate file under `base` with columns matching
    TARGET_COLUMNS. Returns the Path with the highest match score or None
    if no candidate meets MIN_MATCH_SCORE.
    """
    files: list[Path] = []
    for suffix in ("*.csv", "*.xlsx", "*.xls", "*.parquet"):
        files.extend(list(base.rglob(suffix)))
    files = [p for p in files if "venv" not in p.parts and ".git" not in p.parts and not p.name.startswith("~$")]
    best: Path | None = None
    best_score = -1
    for p in files:
        try:
            header = try_read_any(p).head(0)
            found = {normalize_col(c) for c in header.columns}
            score = len(TARGET_COLUMNS.intersection(found))
            if score >= MIN_MATCH_SCORE and score > best_score:
                best, best_score = p, score
        except Exception:
            continue
    return best


def synth_sample_df(n: int = 200) -> pd.DataFrame:
    """
    Generate a synthetic sample dataset with structure similar to the retail
    transactions dataset. Useful when no real data is provided or found.
    """
    rng = np.random.default_rng(42)
    invoice_base = rng.integers(10000, 99999, size=n).astype(str)
    cancel_mask = rng.random(n) < 0.05
    # Initialize invoice array and prefix 'C' to cancelled invoices using vectorized operations
    invoice = invoice_base.copy()
    if cancel_mask.any():
        invoice[cancel_mask] = np.char.add("C", invoice[cancel_mask])
    unit_price = np.round(rng.uniform(0, 50, size=n), 2)
    zero_mask = rng.random(n) < 0.03
    unit_price[zero_mask] = 0.0
    df = pd.DataFrame(
        {
            "InvoiceNo": invoice,
            "StockCode": rng.choice(
                ["10001", "10002", "POST", "D", "M", "20001", "20002", "PADS"],
                size=n,
            ),
            "Description": rng.choice(
                ["MUG", "CUP", "POSTAGE", "Manual", "nan", "check", "PLATE"],
                size=n,
            ),
            "Quantity": rng.integers(1, 8, size=n),
            "InvoiceDate": pd.to_datetime("2021-01-01")
            + pd.to_timedelta(rng.integers(0, 365, size=n), unit="D"),
            "UnitPrice": unit_price,
            "CustomerID": rng.choice(
                [np.nan, 12345, 67890, 11223, 33445],
                size=n,
                p=[0.15, 0.25, 0.2, 0.2, 0.2],
            ),
            "Country": rng.choice(
                ["United Kingdom", "Germany", "France", "Netherlands"],
                size=n,
            ),
        }
    )
    return df


def quality_split(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Split the raw DataFrame into clean and report DataFrames based on the
    number of non-missing values in each row. Rows with fewer than
    VALID_DATA_THRESHOLD valid values are treated as anomalies.
    """
    df = df_raw.copy()
    if "Unnamed: 8" in df.columns:
        df = df.drop(columns=["Unnamed: 8"])
    # Ensure all object columns are strings
    for c in df.select_dtypes(include="object").columns:
        df[c] = df[c].astype(str)
    # Convert date column if present
    if "InvoiceDate" in df.columns:
        df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"], errors="coerce")

    df["valid_values_count"] = df.notna().sum(axis=1)
    conditions = [
        (df["valid_values_count"] == 0),
        (df["valid_values_count"] < VALID_DATA_THRESHOLD),
        (df["valid_values_count"] >= VALID_DATA_THRESHOLD),
    ]
    df["quality_status"] = np.select(
        conditions,
        ["Baris Kosong Total", "Data Sangat Minim", "Data Lengkap"],
        default="Unknown",
    )
    df["OriginalRowNumber"] = np.arange(len(df)) + 2

    df_clean = (
        df[df["quality_status"] == "Data Lengkap"]
        .drop(columns=["valid_values_count", "quality_status", "OriginalRowNumber"])
        .reset_index(drop=True)
    )
    df_report = df[df["quality_status"] != "Data Lengkap"].sort_values(
        by="valid_values_count"
    )
    return df_clean, df_report


# =========================
# CLASSIFICATION
# =========================

def classify_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """
    Label each transaction as Cancelled, Admin/Fee, Zero Price, or Normal Sale.
    """
    d = df.copy()
    d["InvoiceNo"] = d["InvoiceNo"].astype(str)
    admin_codes = [
        "POST",
        "D",
        "MANUAL",
        "C2",
        "M",
        "BANK CHARGES",
        "PADS",
        "DOT",
    ]
    cond_cancelled = d["InvoiceNo"].str.startswith("C", na=False)
    cond_admin = d.get("StockCode", pd.Series(False, index=d.index)).isin(admin_codes)
    cond_zero = (
        pd.to_numeric(d.get("UnitPrice", 0), errors="coerce").fillna(0) == 0
    ) & (~cond_cancelled) & (~cond_admin)
    d["TransactionStatus"] = np.select(
        [cond_cancelled, cond_admin, cond_zero],
        ["Cancelled", "Admin/Fee", "Zero Price"],
        default="Normal Sale",
    )
    return d


# =========================
# VISUALS
# =========================

def plot_treemap(df_classified: pd.DataFrame):
    s = df_classified["TransactionStatus"].value_counts().reset_index()
    s.columns = ["TransactionStatus", "Jumlah"]
    fig = px.treemap(
        s,
        path=["TransactionStatus"],
        values="Jumlah",
        title="Peta Komposisi Status Transaksi",
    )
    fig.update_traces(textinfo="label+value+percent parent")
    fig.update_layout(margin=dict(t=40, l=10, r=10, b=10), height=360)
    return fig


def plot_pie(df_classified: pd.DataFrame):
    s = df_classified["TransactionStatus"].value_counts().reset_index()
    s.columns = ["Status", "Jumlah"]
    fig = px.pie(
        s,
        names="Status",
        values="Jumlah",
        hole=0.4,
        title="Distribusi Status Transaksi (%)",
    )
    fig.update_traces(textinfo="percent+label")
    fig.update_layout(margin=dict(t=40, l=10, r=10, b=10), height=360)
    return fig


def plot_funnel(df_classified: pd.DataFrame, df_report: pd.DataFrame):
    status_counts = df_classified["TransactionStatus"].value_counts()
    labels = ["Data Siap Pakai"] + status_counts.index.tolist()
    values = [len(df_classified)] + status_counts.values.tolist()
    fig = go.Figure(
        go.Funnel(
            y=labels,
            x=values,
            textposition="inside",
            textinfo="value+percent initial",
            hovertemplate="<b>%{y}</b><br>Jumlah: %{x:,}<extra></extra>",
        )
    )
    fig.update_layout(
        title_text="Aliran & Komposisi Data Siap Pakai",
        margin=dict(t=50, l=30, r=30, b=20),
        height=420,
    )
    return fig


# =========================
# NARASI AI (opsional)
# =========================

def _clean_ai_html(s: str) -> str:
    """
    Clean AI-generated HTML: remove code fences and convert **bold** to
    <strong> tags.
    """
    if not s:
        return ""
    s = s.strip()
    # Remove code fences like ```html or ```
    s = s.replace("```html", "").replace("```", "")
    # Replace markdown bold with <strong>
    s = re.sub(r"\*\*(.*?)\*\*", r"<strong>\1</strong>", s)
    return s


def generate_ai_narrative(
    stats: dict, df_classified: pd.DataFrame, top_cancel_item: tuple[str, int], zero_price_count: int
) -> str:
    """
    Generate a short narrative using Gemini if enabled. All numerical statistics
    are precomputed in Python to minimize hallucination and ensure correct
    formatting. The output is cleaned of Markdown and styled for better
    visibility in Streamlit.
    """
    if not ENABLE_AI:
        return ""
    try:
        model = genai.GenerativeModel("gemini-2.5-flash")

        # Precomputed statistics
        status_counts = df_classified["TransactionStatus"].value_counts()
        total = int(status_counts.sum())
        normal_n = int(status_counts.get("Normal Sale", 0))
        cancelled_n = int(status_counts.get("Cancelled", 0))
        zero_n = int(status_counts.get("Zero Price", 0))
        normal_pct = round((normal_n / total * 100) if total else 0.0, 2)
        cancelled_pct = round((cancelled_n / total * 100) if total else 0.0, 2)
        zero_pct = round((zero_n / total * 100) if total else 0.0, 2)
        top_name, top_cnt = top_cancel_item

        RULES = (
            "Gunakan hanya HTML. Tegaskan angka penting dengan <strong>. "
            "Jangan gunakan tanda ** markdown. Hindari tanda hubung panjang."
        )

        # Prompt 1 — Executive Summary
        prompt1 = textwrap.dedent(
            f"""
            {RULES}
            Tulis satu paragraf HTML untuk Ringkasan Eksekutif berdasarkan data:
            - Total transaksi siap pakai: {total:,}
            - Normal Sale: {normal_n:,} ({normal_pct}%)
            - Cancelled: {cancelled_n:,} ({cancelled_pct}%)
            - Zero Price: {zero_n:,} ({zero_pct}%)
            Soroti hanya inti bisnis & risiko utama secara singkat.
            """
        )
        resp1 = model.generate_content(prompt1)
        exec_html = _clean_ai_html(
            resp1.text if hasattr(resp1, "text") else str(resp1)
        )

        # Prompt 2 — Findings & Recommendations
        prompt2 = textwrap.dedent(
            f"""
            {RULES}
            Buat temuan & rekomendasi tindakan dalam HTML (gunakan <h4> dan <p>):
            - Item paling sering dibatalkan: {top_name} ({top_cnt} kali)
            - Zero Price: {zero_price_count:,} transaksi
            Cantumkan 2-3 rekomendasi ringkas & praktis (audit proses, validasi input, pemantauan).
            """
        )
        resp2 = model.generate_content(prompt2)
        find_html = _clean_ai_html(
            resp2.text if hasattr(resp2, "text") else str(resp2)
        )

        # Determine text color from theme palette stored in session
        color = st.session_state.get("ui_text_color")
        if not color:
            try:
                base = st.get_option("theme.base") or "light"
            except Exception:
                base = "light"
            color = "#e5e7eb" if str(base).lower() == "dark" else "#111827"

        # Use consistent color for headings and text
        html = f"""
        <div style='font-family:Inter,Segoe UI; color:{color}; line-height:1.6'>
          <h3 style='margin-top:0; color:{color}'>Analisis Naratif Otomatis</h3>
          <h4 style='color:{color}'>Ringkasan Eksekutif</h4>
          {exec_html}
          <h4 style='color:{color}'>Temuan & Rekomendasi</h4>
          {find_html}
        </div>
        """
        # Final cleanup (remove stray **) if present
        html = html.replace("**", "")
        return html
    except Exception as e:
        logging.warning(f"Gagal narasi AI: {e}")
        return ""


# =========================
# OFFLINE REPORT GENERATOR
# =========================
def build_offline_html_report(
    df_class: pd.DataFrame,
    df_report: pd.DataFrame,
    summary_df: pd.DataFrame,
    narrative_html: str,
    author: str = "Nur Fajril Dwi Pamungkas",
) -> bytes:
    """
    Build a standalone HTML report containing KPI, Plotly charts, summary table, and narrative.
    The resulting file is self-contained and can be opened offline in any browser.

    Args:
        df_class (DataFrame): Clean classified dataset.
        df_report (DataFrame): Report/anomaly dataset.
        summary_df (DataFrame): Summary table of transaction statuses.
        narrative_html (str): Narrative generated (may be empty).
        author (str): Name of the author for attribution.

    Returns:
        bytes: The HTML file contents in UTF-8 encoding.
    """
    # Generate Plotly charts and embed plotlyjs only once for the first figure
    funnel_fig = plot_funnel(df_class, df_report)
    treemap_fig = plot_treemap(df_class)
    pie_fig = plot_pie(df_class)

    # Determine palette based on current theme
    pal = st.session_state.get("ui_palette", get_theme_palette("dark"))
    # Apply template per theme for offline charts
    for fig in [funnel_fig, treemap_fig, pie_fig]:
        try:
            fig.update_layout(template=pal["plotly_template"])
        except Exception:
            pass

    # Convert figures to HTML strings
    funnel_html = funnel_fig.to_html(include_plotlyjs="full", full_html=False)
    treemap_html = treemap_fig.to_html(include_plotlyjs=False, full_html=False)
    pie_html = pie_fig.to_html(include_plotlyjs=False, full_html=False)

    # Prepare summary table HTML
    summary_display = summary_df.reset_index().rename(columns={"index": "Status"})
    table_html = summary_display.to_html(index=False, border=0)

    # KPIs
    total_all = len(df_class) + len(df_report)
    total_clean = len(df_class)
    total_anom = len(df_report)
    now_str = datetime.now().strftime("%d %B %Y, %H:%M WIB")

    # Compose the full HTML with palette
    html_body = f"""
<!doctype html>
<html lang="id">
<head>
<meta charset="utf-8">
<title>Compliance Dashboard — Offline Report</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  body{{background:{pal['bg']};color:{pal['text']};font-family:Inter,Segoe UI,Arial,sans-serif;padding:24px;}}
  h1,h2,h3{{color:{pal['header']};margin:0.2rem 0;}}
  .muted{{opacity:.75;}}
  .kpis{{display:flex;gap:40px;margin:16px 0 24px;}}
  .kpi h3{{margin:0 0 4px 0;font-weight:600;color:{pal['text']};}}
  .kpi div{{font-size:28px;font-weight:700;color:{pal['header']};}}
  .grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;}}
  table{{border-collapse:collapse;width:100%;margin-top:12px;}}
  th,td{{border-bottom:1px solid {pal['border']};padding:8px;text-align:left;color:{pal['text']};}}
  th{{font-weight:600;}}
  .card{{background:{pal['card']};border:1px solid {pal['border']};border-radius:12px;padding:16px;}}
</style>
</head>
<body>
  <h1>Compliance Intelligence Dashboard — Retail</h1>
  <p class="muted">Generated {now_str} • By {author}</p>

  <div class="kpis">
    <div class="kpi"><h3>Total (estimasi mentah)</h3><div>{total_all:,}</div></div>
    <div class="kpi"><h3>Data Siap Pakai</h3><div>{total_clean:,}</div></div>
    <div class="kpi"><h3>Baris Dipisahkan (Anomali)</h3><div>{total_anom:,}</div></div>
  </div>

  <div class="grid">
    <div class="card">{funnel_html}</div>
    <div class="card">{treemap_html}{pie_html}</div>
  </div>

  <div class="card" style="margin-top:16px">
    <h2>Analisis Naratif Otomatis</h2>
    {narrative_html if narrative_html else '<p class="muted">Narasi tidak tersedia.</p>'}
  </div>

  <div class="card" style="margin-top:16px">
    <h2>Ringkasan Status</h2>
    {table_html}
  </div>
</body>
</html>
    """
    return html_body.encode("utf-8")


# =========================
# STREAMLIT UI HELPERS
# =========================

@st.cache_data(show_spinner=False)
def run_pipeline(
    file_bytes: bytes | None, fname: str | None, use_sample: bool
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Execute the full pipeline: loading data, splitting quality, and
    classification. Returns (df_raw, df_clean, df_report, df_classified).
    """
    # 1) Loading
    if file_bytes is not None and fname is not None:
        if fname.lower().endswith(".csv"):
            df_raw = pd.read_csv(io.BytesIO(file_bytes))
        elif fname.lower().endswith((".xlsx", ".xls")):
            df_raw = pd.read_excel(io.BytesIO(file_bytes))
        else:
            raise ValueError("Hanya CSV/XLSX yang diterima dari uploader.")
    elif use_sample:
        # Locate any sample file in the sample_data directory. We prefer
        # compressed CSV (.csv.gz) first, then plain CSV, then Parquet. If
        # nothing is found, fall back to a synthetic dataset.
        sample_path = locate_sample()
        if sample_path is not None:
            try:
                df_raw = try_read_any(sample_path)
            except Exception:
                # If reading fails (e.g. unsupported format), fall back to synthetic sample
                df_raw = synth_sample_df(400)
        else:
            df_raw = synth_sample_df(400)
    else:
        cand = guess_best_candidate(Path.cwd())
        df_raw = try_read_any(cand) if cand else synth_sample_df(400)

    # 2) Quality split
    df_clean, df_report = quality_split(df_raw)

    # 3) Classification
    df_class = classify_transactions(df_clean)

    return df_raw, df_clean, df_report, df_class


def show_pipeline_results(df_class: pd.DataFrame, df_report: pd.DataFrame, save_results: bool = False) -> None:
    """
    Render all dashboard components given classified and report DataFrames.
    If save_results=True, the results will also be written to the persistence
    files.
    """
    # Save results for persistence if requested
    if save_results:
        try:
            # When Parquet support is available, use Parquet; otherwise fall back to gzip-compressed CSV
            if HAVE_PARQUET:
                df_class.to_parquet(LAST_CLASSIFIED_PATH)
                df_report.to_parquet(LAST_REPORT_PATH)
            else:
                df_class.to_csv(LAST_CLASSIFIED_CSV, index=False, compression='gzip')
                df_report.to_csv(LAST_REPORT_CSV, index=False, compression='gzip')
        except Exception as e:
            logging.warning(f"Gagal menyimpan hasil ke file: {e}")

    # Initialize narrative HTML placeholder
    narrative_html: str = ""

    # Metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(
            "Total (estimasi mentah)", value=f"{len(df_class) + len(df_report):,}"
        )
    with col2:
        st.metric("Data Siap Pakai", value=f"{len(df_class):,}")
    with col3:
        st.metric("Baris Dipisahkan (Anomali)", value=f"{len(df_report):,}")
    st.divider()

    # Visuals
    c1, c2 = st.columns([2, 1])
    with c1:
        st.plotly_chart(
            plot_funnel(df_class, df_report), use_container_width=True
        )
    with c2:
        st.plotly_chart(plot_treemap(df_class), use_container_width=True)
        # Hide pie chart behind expander
        with st.expander(
            "📊 Tampilkan Pie Chart Distribusi (%)", expanded=False
        ):
            st.plotly_chart(plot_pie(df_class), use_container_width=True)

    # Summary table
    st.subheader("Ringkasan Status Transaksi")
    status_counts = df_class["TransactionStatus"].value_counts()
    summary_df = pd.DataFrame(
        {
            "Jumlah Transaksi": status_counts,
            "Persentase (%)": (status_counts / len(df_class) * 100).round(2),
        }
    ).sort_values(by="Jumlah Transaksi", ascending=False)
    st.dataframe(summary_df, use_container_width=True)

    # Business insights
    st.subheader("Insight Bisnis")
    # Determine top cancelled item (excluding administrative codes/descriptions)
    non_product_codes = [
        "POST",
        "D",
        "MANUAL",
        "C2",
        "M",
        "BANK CHARGES",
        "PADS",
        "DOT",
    ]
    non_product_descs = ["nan", "?", "check", "POSTAGE", "Manual"]
    cancelled_items_df = df_class[
        (df_class["TransactionStatus"] == "Cancelled")
        & (~df_class["StockCode"].isin(non_product_codes))
        & (~df_class["Description"].isin(non_product_descs))
    ]
    top_cancel = cancelled_items_df["Description"].value_counts().nlargest(1)
    top_name = top_cancel.index[0] if not top_cancel.empty else "N/A"
    top_cnt = int(top_cancel.iloc[0]) if not top_cancel.empty else 0
    zero_n = int((df_class["TransactionStatus"] == "Zero Price").sum())
    normal_n = int((df_class["TransactionStatus"] == "Normal Sale").sum())

    # AI narrative: generate and display if available
    if ENABLE_AI:
        html_ai = generate_ai_narrative(
            {"clean": len(df_class), "anomaly": len(df_report)},
            df_class,
            (top_name, top_cnt),
            zero_n,
        )
        if html_ai:
            # Display AI-generated narrative and store for offline report
            # Reduce the height to fit better on mobile screens
            st.components.v1.html(html_ai, height=520, scrolling=True)
            narrative_html = html_ai
        else:
            st.info(
                "Narasi AI tidak tersedia saat ini. Menampilkan ringkasan non-AI."
            )

    # Non-AI summary (baseline) rendered with HTML to leverage CSS variables.  This
    # approach ensures the text remains legible across light and dark themes.
    cancelled_n = summary_df.loc['Cancelled', 'Jumlah Transaksi'] if 'Cancelled' in summary_df.index else 0
    cancelled_pct = summary_df.loc['Cancelled', 'Persentase (%)'] if 'Cancelled' in summary_df.index else 0.0
    summary_html = (
        f"<ul style='padding-left:1.2rem; margin-bottom:0.5rem;'>"
        f"<li><strong style='color:var(--ui-header-color);'>Cancelled</strong>: {cancelled_n:,} transaksi ({cancelled_pct:.2f}%)."
        f" Item paling sering dibatalkan: <strong style='color:var(--ui-header-color);'>{top_name}</strong> ({top_cnt}×).</li>"
        f"<li><strong style='color:var(--ui-header-color);'>Zero Price</strong>: {zero_n:,} transaksi tanpa pendapatan → audit alur input harga &amp; validasi otomatis.</li>"
        f"<li><strong style='color:var(--ui-header-color);'>Normal Sale</strong>: {normal_n:,} transaksi → basis bisnis sehat; pertahankan keandalan proses.</li>"
        "</ul>"
    )
    st.markdown(summary_html, unsafe_allow_html=True)

    # Data preview with a labeled index column
    st.subheader("Pratinjau Data")
    tab1, tab2, tab3, tab4 = st.tabs(
        ["Clean", "Anomali", "Zero Price", "Cancelled"]
    )
    with tab1:
        df_preview = df_class.head(300).copy()
        df_preview.index = df_preview.index + 1  # make index 1-based
        df_preview.index.name = "Baris"
        st.dataframe(df_preview, use_container_width=True)
    with tab2:
        df_preview = df_report.head(300).copy()
        df_preview.index = df_preview.index + 1
        df_preview.index.name = "Baris"
        st.dataframe(df_preview, use_container_width=True)
    with tab3:
        df_zp = df_class[df_class["TransactionStatus"] == "Zero Price"].head(300).copy()
        df_zp.index = df_zp.index + 1
        df_zp.index.name = "Baris"
        st.dataframe(df_zp, use_container_width=True)
    with tab4:
        df_cancelled = df_class[df_class["TransactionStatus"] == "Cancelled"].head(300).copy()
        df_cancelled.index = df_cancelled.index + 1
        df_cancelled.index.name = "Baris"
        st.dataframe(df_cancelled, use_container_width=True)

    # Downloads
    st.subheader("Unduh Laporan (.xlsx)")
    now_id = datetime.now().strftime("%Y%m%d_%H%M")

    def make_btn(df: pd.DataFrame, title: str, cat: str, fname_prefix: str) -> None:
        bytes_xlsx = df_to_excel_bytes(
            df,
            title,
            {
                "Tanggal Laporan": datetime.now().strftime("%d %B %Y, %H:%M WIB"),
                "Total Baris": f"{len(df):,}",
                "Kategori": cat,
            },
        )
        st.download_button(
            label=f"⬇️ {title}",
            data=bytes_xlsx,
            file_name=f"{fname_prefix}_{now_id}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    cdl, cdr = st.columns(2)
    with cdl:
        make_btn(
            df_class,
            "Laporan Lengkap (Clean)",
            "Data Siap Pakai",
            "laporan_kepatuhan_lengkap",
        )
        make_btn(
            df_class[df_class["TransactionStatus"] == "Zero Price"],
            "Zero Price",
            "Transaksi Harga Nol",
            "zero_price",
        )
        make_btn(
            df_report,
            "Seluruh Anomali",
            "Data Anomali",
            "anomali",
        )
    with cdr:
        for st_name in df_class["TransactionStatus"].unique():
            make_btn(
                df_class[df_class["TransactionStatus"] == st_name],
                f"Status: {st_name}",
                f"Status {st_name}",
                f"status_{st_name.replace(' ', '_')}",
            )

    st.success(
        "✅ Pipeline selesai."
    )

    # Offline HTML report
    # Generate offline report bytes only when needed. Summary_df computed earlier.
    try:
        offline_html = build_offline_html_report(
            df_class,
            df_report,
            summary_df,
            narrative_html,
            author="Nur Fajril Dwi Pamungkas",
        )
        st.download_button(
            "🗂️ Simpan Laporan Offline (.html)",
            data=offline_html,
            file_name=f"retail_dashboard_{now_id}.html",
            mime="text/html",
        )
    except Exception as e:
        logging.warning(f"Gagal membuat laporan offline: {e}")


# =========================
# MAIN UI LOGIC
# =========================


with st.sidebar:
    # Theme selector
    st.header("🎨 Theme")
    _choice = st.radio("Tampilan", ["Auto", "Light", "Dark"], index=0, horizontal=True)
    st.session_state[THEME_KEY] = _choice.lower() if _choice != "Auto" else "auto"

    st.header("⚙️ Input Data")
    up = st.file_uploader(
        "Upload CSV / XLSX (opsional)", type=["csv", "xlsx", "xls"]
    )
    # Pengaturan default: selalu gunakan dataset contoh jika tidak ada file di-upload
    # dan selalu aktifkan narasi AI (demo). Checkbox dihilangkan agar tidak perlu
    # diklik oleh pengguna.
    # Set flag use_ai di session_state ke True secara permanen.
    st.session_state["use_ai"] = True
    # Beri penjelasan singkat bahwa data contoh akan dipakai bila upload kosong
    st.write(
        "Jika Anda tidak mengunggah file, aplikasi secara otomatis akan memproses dataset contoh."
    )
    run_btn = st.button("🚀 Jalankan Pipeline")

# Apply the theme settings after sidebar inputs are defined
apply_theme()

# Render the page title and caption using dynamic CSS variables for color.  The caption
# sits directly beneath the title and leverages the variables defined in apply_theme().
st.title("📊 Compliance Intelligence Dashboard — Retail Transactions")
_caption_text = (
    "Demo publik: pipeline kualitas data → klasifikasi transaksi → insight bisnis + unduhan Excel "
    "— by Nur Fajril Dwi Pamungkas"
)
st.markdown(
    f"<div style='font-size:0.85rem; margin-top:-0.4rem; color: var(--ui-text-color); opacity:0.75'>{_caption_text}</div>",
    unsafe_allow_html=True,
)


def load_or_generate_initial_results() -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Attempt to load previously saved results from disk. If files do not
    exist or fail to load, run the pipeline on a sample or synthetic dataset,
    save results (using Parquet if available, otherwise gzip-compressed CSV),
    and return them.
    """
    # Try to load previously persisted results. Prefer Parquet if available
    try:
        if HAVE_PARQUET and os.path.exists(LAST_CLASSIFIED_PATH) and os.path.exists(LAST_REPORT_PATH):
            df_class = pd.read_parquet(LAST_CLASSIFIED_PATH)
            df_report = pd.read_parquet(LAST_REPORT_PATH)
            return df_class, df_report
        # Fallback to CSV persistence when Parquet is unavailable or not found
        if os.path.exists(LAST_CLASSIFIED_CSV) and os.path.exists(LAST_REPORT_CSV):
            df_class = pd.read_csv(LAST_CLASSIFIED_CSV)
            df_report = pd.read_csv(LAST_REPORT_CSV)
            return df_class, df_report
    except Exception as e:
        logging.warning(
            f"Gagal memuat file hasil terakhir: {e}. Menjalankan pipeline default."
        )

    # No previous files or failed to load; run default pipeline
    # Locate any sample file from sample_data. Use compressed CSV first, then CSV, then Parquet
    sample_path = locate_sample()
    if sample_path is not None:
        try:
            df_raw = try_read_any(sample_path)
        except Exception:
            df_raw = synth_sample_df(400)
    else:
        df_raw = synth_sample_df(400)

    df_clean, df_report = quality_split(df_raw)
    df_class = classify_transactions(df_clean)
    # Save results for persistence
    try:
        if HAVE_PARQUET:
            df_class.to_parquet(LAST_CLASSIFIED_PATH)
            df_report.to_parquet(LAST_REPORT_PATH)
        else:
            df_class.to_csv(LAST_CLASSIFIED_CSV, index=False, compression='gzip')
            df_report.to_csv(LAST_REPORT_CSV, index=False, compression='gzip')
    except Exception as e:
        logging.warning(f"Gagal menyimpan hasil default: {e}")
    return df_class, df_report


if run_btn:
    # User requested to run the pipeline
    with st.spinner("Memproses data..."):
        file_bytes, fname = (up.read(), up.name) if up is not None else (None, None)
        # Jalankan pipeline selalu dengan parameter use_sample=True. Jika file diunggah,
        # pipeline akan memakai file tersebut; jika tidak, pipeline akan memuat
        # dataset contoh dari sample_data atau membuat data sintetis.
        df_raw, df_clean, df_report_new, df_class_new = run_pipeline(
            file_bytes, fname, True
        )
    # Display results and save them
    show_pipeline_results(df_class_new, df_report_new, save_results=True)
else:
    # Initial load: show previous or default results
    df_class_prev, df_report_prev = load_or_generate_initial_results()
    show_pipeline_results(df_class_prev, df_report_prev, save_results=False)
